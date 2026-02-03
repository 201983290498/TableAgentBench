"""
Collection of table processing utility functions

Includes the following features:
- format_table_desc: Format table description information (dict format) into text
- read_table_lines/read_all_sheets_lines: Read table row data
- get_table_preview_str/get_all_sheets_preview: Generate table preview
- convert_xls_sheet_to_rows/convert_sheet_to_csv: Table conversion
"""
import os
import csv
import re
from typing import List, Dict, Any, Union, Optional, Tuple
from pathlib import Path
from src.utils.common import read_text_file
from openpyxl import load_workbook
try:
    import xlrd
except ImportError as e:
    print(f"Warning: Failed to import xlrd: {e}")
    xlrd = None
import csv
import numpy as np

def read_table_lines(file_path: str, head_n: int = 5, tail_n: int = 5, mode: str = "default") -> Tuple[List[str], List[str], int]:
    """
    Read the first N rows and last N rows of a file as plain text.
    Supports csv, xlsx, xls, tsv file formats (xlsx/xls will be converted to temporary CSV first).
    
    Args:
        file_path: Path to the file
        head_n: Number of first N rows
        tail_n: Number of last N rows
        mode: Reading mode
            - "default": Returns first N and last N rows
            - "all": Returns all rows (first parameter is all rows, second is an empty list)
        
    Returns:
        (head_lines, tail_lines, total_lines) tuple
        - mode="default": (First N rows, Last N rows, Total count)
        - mode="all": (All rows, [], Total count)
    """
    assert os.path.isfile(file_path), f"File does not exist: {file_path}"
    file_ext = os.path.splitext(file_path)[1].lower()
    # Excel files: try openpyxl first, then xlrd (compatible with various formats)
    if file_ext in ['.xlsx', '.xls']:
        try:
            # Attempt 1: openpyxl (mainly for xlsx)
            try:
                # Open with file object to bypass openpyxl's .xls extension check
                with open(file_path, 'rb') as f:
                    wb = load_workbook(f, data_only=True)
                    if len(wb.sheetnames) > 1:
                        print(f"Warning: File {file_path} contains multiple sheets, only processing the first sheet: {wb.sheetnames[0]}")
                    result = convert_sheet_to_csv(wb[wb.sheetnames[0]])
                    wb.close()
                    lines = result.get('rows_data', [])
            except Exception as e_openpyxl:
                # Attempt 2: xlrd (for xls and potential old formats)
                # Note: xlrd 2.0+ does not support xlsx and will throw XLRDError
                wb = xlrd.open_workbook(file_path)
                if len(wb.sheet_names()) > 1:
                    print(f"Warning: File {file_path} contains multiple sheets, only processing the first sheet: {wb.sheet_names()[0]}")
                result = convert_xls_sheet_to_rows(wb.sheet_by_index(0))
                lines = result.get('rows_data', [])
        except Exception as e:
            # All attempts failed, print error and return empty
            print(f"Failed to read table file {file_path}: {str(e)}")
            return [], [], 0
    else:
        # Text formats like csv, tsv
        lines = [line.rstrip('\n\r') for line in read_text_file(file_path)]
    # Return different results based on mode
    if mode == "all":
        return lines, [], len(lines)
    tail_lines = lines[-tail_n:] if tail_n > 0 else []
    return lines[:head_n], tail_lines, len(lines)


def read_all_sheets_lines(file_path: str, sheet_name: Optional[str] = None) -> List[Tuple[str, List[str]]]:
    """
    Read all contents of a file, supporting multiple sheets.
    Supports csv, xlsx, xls, tsv file formats.
    
    Args:
        file_path: Path to the file
        sheet_name: Specified sheet name to read, default is None (reads all)
        
    Returns:
        List[Tuple[str, List[str]]]: 
        Each element in the list is (name, list of row data)
        - For single sheet or text files: name is the filename
        - For multiple sheets: name is "filename::sheetname"
    """
    assert os.path.isfile(file_path), f"File does not exist: {file_path}"
    file_ext = os.path.splitext(file_path)[1].lower()
    file_name = os.path.basename(file_path)
    results = []
    # Excel files: try openpyxl first, then xlrd (compatible with various formats)
    if file_ext in ['.xlsx', '.xls']:
        try:
            # Attempt 1: openpyxl (mainly for xlsx)
            try:
                # Open with file object to bypass openpyxl's .xls extension check
                with open(file_path, 'rb') as f:
                    wb = load_workbook(f, data_only=True)
                    sheet_names = wb.sheetnames
                    # If sheet_name is specified, check if it exists
                    if sheet_name:
                        if sheet_name not in sheet_names:
                            print(f"Warning: Specified sheet {sheet_name} not found in file {file_path}")
                            return []
                        target_sheets = [sheet_name]
                    else:
                        target_sheets = sheet_names
                    use_sheet_name = len(target_sheets) > 1 or (len(sheet_names) > 1)
                    for s_name in target_sheets:
                        result = convert_sheet_to_csv(wb[s_name])
                        lines = result.get('rows_data', [])
                        if use_sheet_name:
                            name = f"{file_name}::{s_name}"
                        else:
                            name = file_name
                        results.append((name, lines))
                    wb.close()
                    return results
            except Exception as e_openpyxl:
                # Attempt 2: xlrd (for xls and potential old formats)
                wb = xlrd.open_workbook(file_path)
                sheet_names = wb.sheet_names()
                
                # If sheet_name is specified, check if it exists
                if sheet_name:
                    if sheet_name not in sheet_names:
                        print(f"Warning: Specified sheet {sheet_name} not found in file {file_path}")
                        return []
                    target_sheets = [sheet_name]
                else:
                    target_sheets = sheet_names
                    
                use_sheet_name = len(target_sheets) > 1 or (len(sheet_names) > 1)
                for s_name in target_sheets:
                    try:
                        sheet = wb.sheet_by_name(s_name)
                        result = convert_xls_sheet_to_rows(sheet)
                        lines = result.get('rows_data', [])
                        if use_sheet_name:
                            name = f"{file_name}::{s_name}"
                        else:
                            name = file_name
                        results.append((name, lines))
                    except:
                        continue
                return results
        except Exception as e:
            # All attempts failed, print error and return empty
            print(f"Failed to read table file {file_path}: {str(e)}")
            return []
    else:
        # Text formats like csv, tsv
        lines = [line.rstrip('\n\r') for line in read_text_file(file_path)]
        return [(file_name, lines)]


def format_table_desc(table_info: Union[str, Dict, List]) -> str:
    """
    Format table description information (dict format) into text.
    
    Used to convert table metadata (e.g., file path, table description, column info) into readable text.
    Supports a single dict, a list of multiple dicts, or returning a string directly.
    
    Args:
        table_info: Table information, can be:
            - str: Returned directly
            - Dict: Contains file_path, table_desc, column_info fields
            - List: A list of multiple table information entries
            
    Returns:
        Formatted table description text
    """
    if isinstance(table_info, str):
        return table_info
    if isinstance(table_info, list):
        return "\n".join([format_table_desc(t) for t in table_info])
    # dict format
    lines = []
    if 'file_path' in table_info:
        lines.append(f"File path: {table_info['file_path']}")
    if 'table_desc' in table_info:
        lines.append(f"Table description: {table_info['table_desc']}")
    if 'column_info' in table_info:
        lines.append("Column information:")
        for col in table_info['column_info']:
            lines.append(f"  - {col.get('col_name', 'Unknown')} ({col.get('col_type', '')}): {col.get('col_desc', '')}")
    return "\n".join(lines)


def scan_table_files(folder_path: str, extensions: List[str] = None) -> List[Dict[str, Any]]:
    """
    Scan for table files in a folder and retrieve all table-related information.
    
    Args:
        folder_path: Folder path
        extensions: Supported file extensions
        
    Returns:
        List of table file information
    """
    if extensions is None:
        extensions = ['.csv', '.xlsx', '.xls', '.tsv']
    folder = Path(folder_path)
    if not folder.exists():
        return []
    table_files = []
    
    # Recursive scan
    for ext in extensions:
        for file_path in folder.rglob(f'*{ext}'):
            table_files.append({'path': str(file_path), 'name': file_path.name, 'relative_path': str(file_path.relative_to(folder)), 
                                'extension': ext, 'size': file_path.stat().st_size})
    
    return table_files

def get_table_preview_str(file_path: str, start: int = 1, n: int = 5) -> str:
    """
    Get a formatted table preview string (including column names and first few rows of data), 
    similar to TableHeadReader.format_output format.
    
    Args:
        file_path: File path
        start: Starting row (starts from 1)
        n: Number of rows to read
        
    Returns:
        Formatted preview string
    """
    path = Path(file_path)
    if not path.exists():
        return f"[Error] File not found: {path.name}"
    try:
        head_lines, _, total_lines = read_table_lines(str(path), head_n=n, tail_n=0)
        content_lines = []
        for i, line in enumerate(head_lines, start=start):
            content_lines.append(f"{i:3d}| {line.rstrip()}")
        
        shown_count = len(head_lines)
        remaining_count = max(0, total_lines - shown_count)
        
        return (
            f"[{path.name}] Total rows: {total_lines}, remaining rows not shown: {remaining_count}\n"
            f"{'─' * 40}\n"
            + "\n".join(content_lines)
        )
    except Exception as e:
        return f"[Error] Failed to read {path.name}: {str(e)}"

def get_all_sheets_preview(file_path: str, max_rows: int = 3) -> List[Dict[str, Any]]:
    """
    Get preview information for all sheets in a table (column names + first few rows). 
    Supports table files in csv, xlsx, and xls formats.
    
    Args:
        file_path: File path
        max_rows: Maximum number of preview rows
        
    Returns:
        List of preview information dictionaries, each containing file, path, sheet, head_lines, total_rows, error (if any)
    """
    path = Path(file_path)
    previews = []
    try:
        # Read all sheets
        sheets_data = read_all_sheets_lines(file_path)
        for source_name, lines in sheets_data:
            if "::" in source_name:
                # Multi-sheet case
                file_name, sheet_name = source_name.split("::", 1)
            else:
                # Single-sheet case
                file_name = source_name
                sheet_name = ""
            # Extract the first max_rows as preview
            head_lines = lines[:max_rows]
            total_rows = max(0, len(lines) - 1)  # Subtract header row
            preview = {'file': file_name, 'path': str(path), 'sheet': sheet_name, 'head_lines': head_lines, 'total_rows': total_rows}
            previews.append(preview)
    except Exception as e:
        previews.append({'file': path.name, 'path': str(path), 'sheet': "", 'error': str(e)})
    return previews

def _normalize_cell_array(values: np.ndarray) -> np.ndarray:
    """Use numpy vectorized operations to normalize cell values in bulk"""
    # Vectorized processing function
    def process_cell(val):
        if val is None or val == "":
            return ""
        # Return numeric types directly
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return val
        # Handle string types
        if not isinstance(val, str):
            return val
        text = str(val).strip()
        if not text:
            return ""
        # Handle percentage
        percentage_flag = text.endswith('%')
        clean_text = text.replace(',', '').replace('%', '')
        # Try converting to number
        try:
            num = int(clean_text)
            return num / 100 if percentage_flag else num
        except ValueError:
            try:
                num = float(clean_text)
                return num / 100 if percentage_flag else num
            except ValueError:
                return val
    # Use numpy.vectorize for bulk processing
    vectorized_process = np.vectorize(process_cell, otypes=[object])
    return vectorized_process(values)

def _is_junk_val(val: str) -> bool:
    """Check if a value is considered 'junk' (empty or Unnamed: x)"""
    if val is None:
        return True
    s = str(val).strip()
    if s == "":
        return True
    # Check for Unnamed: x
    if re.match(r'^Unnamed:\s*\d+$', s, re.IGNORECASE):
        return True
    return False

def _clean_rows_data(rows: List[List[str]]) -> List[List[str]]:
    """
    Clean rows data by removing:
    1. Leading rows that are all junk (empty or Unnamed: x)
    2. Any columns that are all junk
    """
    if not rows:
        return rows
    
    # 1. Trim leading junk rows
    start_row = 0
    while start_row < len(rows):
        row = rows[start_row]
        if all(_is_junk_val(x) for x in row):
            start_row += 1
        else:
            break
            
    rows = rows[start_row:]
    
    if not rows:
        return []

    # 2. Trim ANY junk columns
    # We need to check if column j is junk for ALL remaining rows
    if not rows:
        return []
        
    num_cols = len(rows[0])
    valid_col_indices = []
    
    for col_idx in range(num_cols):
        is_col_junk = True
        for row in rows:
            # Handle potential uneven rows (though unexpected here)
            val = row[col_idx] if col_idx < len(row) else None
            if not _is_junk_val(val):
                is_col_junk = False
                break
        
        if not is_col_junk:
            valid_col_indices.append(col_idx)
            
    # Reconstruct rows with only valid columns
    cleaned_rows = []
    for row in rows:
        new_row = [row[i] for i in valid_col_indices if i < len(row)]
        cleaned_rows.append(new_row)
            
    return cleaned_rows

def convert_xls_sheet_to_rows(sheet_or_path, csv_path: str = None, encoding: str = "utf-8-sig") -> Dict[str, Any]:
    """
    Convert xlrd sheet object or .xls file to row data
    
    Args:
        sheet_or_path: xlrd sheet object or .xls file path
        csv_path: Output csv path (optional)
        encoding: Encoding format
        
    Returns:
        Transformation result dictionary
    """
    # If a file path string is passed
    if isinstance(sheet_or_path, str):
        wb = xlrd.open_workbook(sheet_or_path)
        sheet = wb.sheet_by_index(0)
        datemode = wb.datemode
    else:
        sheet = sheet_or_path
        datemode = sheet.book.datemode
    
    try:
        # 1. Skip completely empty sheets
        if sheet.nrows == 0 or sheet.ncols == 0:
            return {"success": False, "rows": 0, "cols": 0, "merged_cells": 0, "rows_data": []}
        
        # 2. Read directly into a numpy array (bulk operation, read by line)
        arr = np.empty((sheet.nrows, sheet.ncols), dtype=object)
        cell_types = np.empty((sheet.nrows, sheet.ncols), dtype=int)
        
        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)  # Get the entire row at once
            arr[row_idx] = [cell.value for cell in row]
            cell_types[row_idx] = [cell.ctype for cell in row]
        
        # 3. Fill merged cells in bulk (operate on the numpy array)
        merged_cells_count = 0
        if hasattr(sheet, 'merged_cells'):
            for (min_row, max_row, min_col, max_col) in sheet.merged_cells:
                merged_cells_count += 1
                value = arr[min_row, min_col]  # Get the value of the top-left cell
                arr[min_row:max_row, min_col:max_col] = value  # numpy slice bulk assignment
        
        # 4. Handle date types in bulk (operate on the numpy array) - optimized version
        date_mask = (cell_types == xlrd.XL_CELL_DATE) & (arr != None) & (arr != "")
        if date_mask.any():
            date_indices = np.argwhere(date_mask)
            for row_idx, col_idx in date_indices:
                date_tuple = xlrd.xldate_as_tuple(arr[row_idx, col_idx], datemode)
                arr[row_idx, col_idx] = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
        
        # 5. Normalize all cell values in bulk (numpy vectorization)
        arr = _normalize_cell_array(arr)
        
        # 6. Convert to strings and prepare for output (optimization: reduce intermediate conversions)
        # Convert directly to string matrix to avoid tolist then traversal
        str_arr = np.vectorize(str)(arr)
        all_rows = str_arr.tolist()  # Translate to list once
        
        # 7. Write to CSV file (optimization: let csv.writer handle separators)
        if csv_path:
            with open(csv_path, 'w', encoding=encoding, newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(all_rows)  # Pass list of lists
        
        # 8. Join only when returning (only when string format is needed)
        rows_data = [','.join(row) for row in all_rows]
        return {"success": True, "rows": sheet.nrows, "cols": sheet.ncols, "merged_cells": merged_cells_count, "rows_data": rows_data}
    except Exception as e:
        return {"success": False, "error": str(e), "rows": 0, "cols": 0, "merged_cells": 0, "rows_data": []}

def convert_sheet_to_csv(worksheet, csv_path: str=None, encoding: str = "utf-8-sig") -> Dict[str, Any]:
    """
    Convert a single worksheet to a csv file.
    
    Args:
        worksheet: openpyxl worksheet object or xlsx file path
        csv_path: Output csv path
        encoding: Encoding format
        
    Returns:
        Transformation result dictionary
    """
    if isinstance(worksheet, str):
        wb = load_workbook(worksheet, data_only=True)
        worksheet = wb[wb.sheetnames[0]]
    
    try:
        # 1. Skip completely empty sheets
        if worksheet.max_column == 0 or worksheet.max_row == 0:
            return {"success": False, "rows": 0, "cols": 0, "merged_cells": 0, "rows_data": []}
        
        # 2. Read directly into a numpy array (bulk operation)
        arr = np.empty((worksheet.max_row, worksheet.max_column), dtype=object)
        row_idx = 0
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column, values_only=True):
            arr[row_idx] = row
            row_idx += 1
        # 3. Fill merged cells in bulk (operate on the numpy array)
        merged_cells_count = 0
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells_count += 1
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = arr[min_row - 1, min_col - 1]  # Get the value of the top-left cell
            arr[min_row - 1:max_row, min_col - 1:max_col] = value  # numpy slice bulk assignment
        # 4. Normalize all cell values in bulk (numpy vectorization)
        arr = _normalize_cell_array(arr)
        
        # 5. Convert to strings and prepare for output (optimization: reduce intermediate conversions)
        str_arr = np.vectorize(str)(arr)
        all_rows = str_arr.tolist()
        
        # Clean data: remove invalid leading rows and columns
        all_rows = _clean_rows_data(all_rows)
        
        # 6. Write to CSV file (optimization: let csv.writer handle separators)
        if csv_path:
            with open(csv_path, 'w', encoding=encoding, newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(all_rows)
        
        # 7. Join only when returning
        rows_data = [','.join(row) for row in all_rows]
        return {"success": True, "rows": len(all_rows), "cols": len(all_rows[0]) if all_rows else 0, "merged_cells": merged_cells_count, "rows_data": rows_data}
    except Exception as e:
        return {"success": False, "error": str(e), "rows": 0, "cols": 0, "merged_cells": 0, "rows_data": []}

