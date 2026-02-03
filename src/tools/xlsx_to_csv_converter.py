"""
XLSX to CSV Conversion Tool
Supports splitting multi-sheet xlsx files into multiple csv files and handles merged cells
"""
from typing import List, Dict, Any, Optional
from pathlib import Path
import os
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
try:
    import xlrd
except ImportError:
    xlrd = None

from src.tools.base import BaseTool, ToolResult, register_tool
from src.utils.table_process import convert_sheet_to_csv, convert_xls_sheet_to_rows

@register_tool()
class XlsxToCsvConverter(BaseTool):
    """
    XLSX to CSV Conversion Tool
    
    Features:
    1. Split multi-sheet xlsx files into multiple csv files
    2. Handle merged cells (copy merged cell value to all cells)
    3. Output file naming: {xlsx_filename}_{sheet_name}.csv
    4. Option to convert all sheets or specify certain sheets
    """
    
    name = "xlsx_to_csv_converter"
    description = """Type: Table Processing. Converts xlsx files to csv files, supporting multi-sheet splitting and merged cell filling. It can split an xlsx file containing multiple sheets into multiple csv files, returning the paths of the split csv files and related local preview information.
**Applicable Scenarios**:
1. Suitable as a preprocessing step for complex, ultra-long table analysis.
2. Suitable for decomposing xlsx into multiple intuitive csv files.
"""
    category = "table_process"
    parameters = {
        "xlsx_path": {
            "type": "string",
            "description": "Absolute path of the xlsx file",
            "required": True
        },
        # "output_dir": {
        #     "type": "string",
        #     "description": "Absolute path of output directory, if empty defaults to xlsx file directory",
        #     "required": False
        # },
        "sheet_names": {
            "type": "array",
            "description": "List of sheet names to convert, if empty converts all sheets",
            "items": {        # For array type, items field must define element type
                "type": "string"
            },
            "required": False
        }
    }
    
    def execute(
        self, 
        xlsx_path: str,
        output_dir: str = None,
        sheet_names: List[str] = None,
    ) -> ToolResult:
        """
        Execute xlsx to csv conversion
        
        Args:
            xlsx_path: xlsx file path
            output_dir: Output directory, defaults to xlsx directory
            sheet_names: List of sheets to convert, None means all
            encoding: Output encoding, default utf-8-sig
            
        Returns:
            ToolResult: Contains conversion result information
        """
        try:
            # 1. Validate input file
            # Convert to absolute path to avoid path check issues
            xlsx_path = str(Path(os.path.abspath(xlsx_path)))
            if not os.path.exists(xlsx_path):
                msg = f"File does not exist: {xlsx_path}"
                return ToolResult(success=False, data=msg, message=msg)
            if not xlsx_path.lower().endswith(('.xlsx', '.xls')):
                msg = "File format not supported, only .xlsx and .xls files are supported"
                return ToolResult(success=False, data=msg, message=msg)
            
            # 2. Determine output directory
            xlsx_file = Path(xlsx_path)
            output_dir = output_dir or str(xlsx_file.parent)
            output_path = Path(os.path.abspath(output_dir))
            output_path.mkdir(parents=True, exist_ok=True)
            
            # 3. Load workbook, determine sheets to convert
            file_ext = xlsx_file.suffix.lower()
            if file_ext == '.xlsx':
                wb = load_workbook(xlsx_path, data_only=True)
                available_sheets = wb.sheetnames
            elif file_ext == '.xls':
                wb_xls = xlrd.open_workbook(xlsx_path)
                available_sheets = wb_xls.sheet_names()
            else:
                msg = f"Unsupported file format: {file_ext}"
                return ToolResult(success=False, data=msg, message=msg)
            
            if sheet_names:
                invalid_sheets = [s for s in sheet_names if s not in available_sheets]
                if invalid_sheets:
                    msg = f"The following sheets do not exist: {invalid_sheets}. Available sheets: {available_sheets}"
                    return ToolResult(success=False, data=msg, message=msg)
            sheets_to_convert = sheet_names or available_sheets
            
            # 4. Convert each sheet
            converted_files, conversion_details = [], []
            for sheet_name in sheets_to_convert:
                safe_sheet_name = self._sanitize_filename(sheet_name)
                csv_path = str(output_path / f"{xlsx_file.stem}_{safe_sheet_name}.csv")
                if file_ext == '.xlsx':
                    ws = wb[sheet_name]
                    result = convert_sheet_to_csv(ws, csv_path, "utf-8-sig")
                elif file_ext == '.xls':
                    sheet = wb_xls.sheet_by_name(sheet_name)
                    result = convert_xls_sheet_to_rows(sheet, csv_path, "utf-8-sig")
                if result["success"]:
                    converted_files.append(csv_path)
                    conversion_details.append({
                        "sheet_name": sheet_name,
                        "csv_path": csv_path,
                        **{k: result[k] for k in ["rows", "cols", "merged_cells"]}
                    })
            if file_ext == '.xlsx':
                wb.close()
                
            # 5. Return result
            if not converted_files:
                msg = "No sheets were successfully converted"
                return ToolResult(success=False, data=msg, message=msg)
            summary = self._format_conversion_summary(xlsx_path, converted_files, conversion_details)
            return self.make_result(True, summary, summary)
        except Exception as e:
            msg = f"Conversion failed: {str(e)}"
            return ToolResult(success=False, data=msg, message=msg)
    
    def _sanitize_filename(self, filename: str) -> str:
        """Clean invalid characters from filename"""
        for char in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
            filename = filename.replace(char, '_')
        return filename.strip() or "sheet"
    
    def _format_conversion_summary(self, xlsx_path: str, converted_files: List[str], details: List[Dict[str, Any]]) -> str:
        """Format conversion summary information"""
        from src.utils.table_process import get_table_preview_str

        sheet_list_str = ", ".join([d['sheet_name'] for d in details])
        
        detail_lines = []
        for d in details:
            preview = get_table_preview_str(d['csv_path'], n=10)
            detail_lines.append(
                f"  - Sheet {d['sheet_name']}: {d['rows']} rows x {d['cols']} columns, processed {d['merged_cells']} merged cells\n"
                f"    Output file path: {d['csv_path']}\n"
                f"    Preview:\n{preview}"
            )

        return f"Successfully converted xlsx file to {len(converted_files)} csv files\nSource file: {Path(xlsx_path).name}\nContains Sheets: {sheet_list_str}\n\nConversion details:\n" + "\n\n".join(detail_lines)
